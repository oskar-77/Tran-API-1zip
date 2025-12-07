"""Excel document extractor using openpyxl."""

from pathlib import Path
from typing import Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.extractors.base_extractor import BaseExtractor
from src.schemas.document import (
    Document, Page, Metadata, TextDirection,
    HeadingBlock, TableBlock, TableRow, TableCell
)
from src.utils.text_utils import detect_text_direction, clean_text


class XlsxExtractor(BaseExtractor):
    """Extractor for Excel documents using openpyxl."""
    
    SUPPORTED_EXTENSIONS = ['.xlsx', '.xlsm']
    
    def extract(self) -> Document:
        """Extract content from Excel document."""
        wb = load_workbook(str(self.file_path), data_only=True)
        
        metadata = self._extract_metadata(wb)
        pages = []
        
        for sheet_index, sheet_name in enumerate(wb.sheetnames):
            sheet = wb[sheet_name]
            page = self._extract_sheet(sheet, sheet_index + 1, sheet_name)
            pages.append(page)
        
        all_text = " ".join([
            block.text if hasattr(block, 'text') else 
            " ".join([cell.content for row in block.rows for cell in row.cells])
            for page in pages for block in page.blocks
        ])
        direction = TextDirection(detect_text_direction(all_text))
        
        metadata.page_count = len(pages)
        
        wb.close()
        
        return Document(
            title=metadata.title or self.file_path.stem,
            metadata=metadata,
            pages=pages,
            direction=direction
        )
    
    def _extract_metadata(self, wb) -> Metadata:
        """Extract metadata from Excel workbook."""
        base_metadata = self._create_base_metadata()
        props = wb.properties
        
        return Metadata(
            title=props.title or base_metadata.source_filename,
            author=props.creator,
            created_date=props.created,
            modified_date=props.modified or base_metadata.modified_date,
            subject=props.subject,
            keywords=self._parse_keywords(props.keywords),
            source_format='xlsx',
            source_filename=base_metadata.source_filename
        )
    
    def _parse_keywords(self, keywords_str: Optional[str]) -> list[str]:
        """Parse keywords string into list."""
        if not keywords_str:
            return []
        
        keywords = [k.strip() for k in keywords_str.split(',')]
        return [k for k in keywords if k]
    
    def _extract_sheet(self, sheet: Worksheet, sheet_num: int, sheet_name: str) -> Page:
        """Extract content from a worksheet."""
        blocks = []
        
        blocks.append(HeadingBlock(
            level=2,
            text=sheet_name,
            direction=TextDirection(detect_text_direction(sheet_name)),
            position_hint=f"sheet_{sheet_num}"
        ))
        
        table_block = self._extract_table(sheet, sheet_num)
        if table_block:
            blocks.append(table_block)
        
        all_text = sheet_name
        if table_block:
            all_text += " " + " ".join([
                cell.content for row in table_block.rows for cell in row.cells
            ])
        direction = TextDirection(detect_text_direction(all_text))
        
        return Page(
            page_number=sheet_num,
            title=sheet_name,
            blocks=blocks,
            direction=direction
        )
    
    def _extract_table(self, sheet: Worksheet, sheet_num: int) -> Optional[TableBlock]:
        """Extract table from worksheet."""
        if sheet.max_row is None or sheet.max_column is None:
            return None
        
        if sheet.max_row == 0 or sheet.max_column == 0:
            return None
        
        rows = []
        
        for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 1000))):
            cells = []
            has_content = False
            
            for cell in row:
                cell_value = cell.value
                if cell_value is not None:
                    has_content = True
                    cell_text = clean_text(str(cell_value))
                else:
                    cell_text = ""
                
                cells.append(TableCell(
                    content=cell_text,
                    is_header=(row_index == 0)
                ))
            
            if has_content:
                rows.append(TableRow(
                    cells=cells,
                    is_header_row=(row_index == 0)
                ))
        
        if not rows:
            return None
        
        return TableBlock(
            rows=rows,
            has_header=True,
            row_count=len(rows),
            column_count=len(rows[0].cells) if rows else 0,
            position_hint=f"sheet_{sheet_num}"
        )
